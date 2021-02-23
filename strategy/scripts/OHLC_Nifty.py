# -*- coding: utf-8 -*-
"""
Spyder Editor
Script to get daily OHLC based on the strike price of Index at 09:20 AM everyday.
update only the variable niftyAt920, weekly expiry 

"""
from datetime import datetime,date
import pandas as pd
from XTConnect import XTSConnect
import configparser
from pathlib import Path
from openpyxl import load_workbook
import os.path

cfg = configparser.ConfigParser()
cfg.read('../../XTConnect/config.ini')

source = cfg['user']['source']
appKey = cfg.get('user', 'marketdata_appkey')
secretKey = cfg.get('user', 'marketdata_secretkey')

cdate = datetime.strftime(datetime.now(), "%d-%m-%Y")
xt = XTSConnect(appKey, secretKey, source)

token_file=f'access_token_{cdate}.txt'
file = Path(token_file)
if file.exists() and (date.today() == date.fromtimestamp(file.stat().st_mtime)):
    print('Token file exists and created today')
    in_file = open(token_file,'r').read().split()
    access_token = in_file[0]
    userID=in_file[1]
    # isInvestorClient=in_file[2]
    print('Initializing session with token..')
    xt._set_common_variables(access_token, userID)
# else:
#     print('Creating token file')   
#     response = xt.marketdata_login()
#     print(response)
#     if "token" in response['result']:
#         with open ('access_token.txt','w') as file:
#             file.write('{}\n{}\n'.format(response['result']['token'], response['result']['userID']
#                                            )) 


def strkPrcCalc(spot,base):
    strikePrice = base * round(spot/base)
    # logger.info(f'StrikePrice computed as : {strikePrice}')
    print(f'StrikePrice computed as : {strikePrice}')
    return strikePrice

cdate = datetime.strftime(datetime.now(), "%b %d %Y")
niftyAt920 = 14752
strikePrice = strkPrcCalc(niftyAt920, 50)

if __name__ == '__main__':
    filename=r'..\ohlc\NIFTY_OHLC.xlsx'
    # with pd.ExcelWriter(filename,engine='openpyxl') as writer:
    for i in range(strikePrice-250,strikePrice+300,50):
        print(i)
        for j in ['CE','PE']:
            print(j)
            # try:
            resp=xt.get_option_symbol(
            exchangeSegment=2,
            series='OPTIDX',
            symbol='NIFTY',
            expiryDate='25Feb2021',
            optionType=j,
            strikePrice=i)
            # alist.append([resp['result'][0]['ExchangeInstrumentID'],resp['result'][0]['DisplayName']])
            eid=resp['result'][0]['ExchangeInstrumentID']
            name=resp['result'][0]['Description']
            print(eid, name)
            ohlc = xt.get_ohlc(
            exchangeSegment=xt.EXCHANGE_NSEFO,
            exchangeInstrumentID=eid,
            startTime=cdate+' 091500',
            endTime=cdate+' 153000',
            compressionValue=60)
            # print("OHLC: " + str(ohlc))
            dataresp= ohlc['result']['dataReponse']
            spl = dataresp.split(',')
            spl_df = pd.DataFrame([sub.split("|") for sub in spl],columns=(['Timestamp','Open','High','Low','Close','Volume','OI','NA']))
            spl_df.drop(spl_df.columns[[-1,]], axis=1, inplace=True)
            spl_df['Timestamp'] = pd.to_datetime(spl_df['Timestamp'].astype('int'), unit='s')
            # spl_df.head()
            writer = pd.ExcelWriter(filename,engine='openpyxl')
            writer.book = load_workbook(filename,read_only=False, keep_vba=True)
            writer.sheets=dict((ws.title, ws) for ws in writer.book.worksheets)
            # print("before", writer.sheets.keys())
            if (name+'_'+j) not in list(writer.sheets.keys()):
                print(f"Adding this sheet: {name+'_'+j}")
                writer.book.create_sheet((name+'_'+j))
                writer.sheets=dict((ws.title, ws) for ws in writer.book.worksheets)
                # print("after", writer.sheets.keys())
            startrow = writer.book[(name+'_'+j)].max_row
            spl_df.to_excel(writer, sheet_name=(name+'_'+j), index=False, header=False, startrow=startrow)
            writer.save()
            writer.close()
                # except Exception as e:
                    # print(e)
                    
        print('==========================================')
        # xt.marketdata_logout()

    
