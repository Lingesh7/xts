# -*- coding: utf-8 -*-
"""
Created on Sun Mar  7 13:22:38 2021
Option Scalper Script (Strategy-2)
@author: mling
"""
############## imports ##############
from datetime import datetime,date
from dateutil.relativedelta import relativedelta, TH, WE
from XTConnect.Connect import XTSConnect
import XTConnect.Exception as ex
from pathlib import Path
import time
import json
import logging
import pandas as pd
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
# pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', -1)
import configparser
import argparse
import timer
from threading import Thread
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from logging.handlers import TimedRotatingFileHandler
from sys import exit
import os
from random import randint

try:
    os.chdir(r'D:\Python\First_Choice_Git\xts\strategy\scripts')
except:
    pass
############## parsing args ##############

parser = argparse.ArgumentParser(description='OptionScalper Script')
parser.add_argument('-t', '--ticker',type=str, required=True, help='NIFTY or BANKNIFTY')
parser.add_argument('-st', '--startTime',type=str, required=True, help='start time of the script')
parser.add_argument('-et', '--endTime',type=str, default="15:05:00", help='end time')
parser.add_argument('-rt', '--repairTime',type=str, default="14:40:00", help='reapir time')
parser.add_argument('-sl', '--stopLoss',type=int, default=-1500, help='stopLoss amount')
parser.add_argument('-tgt', '--target',type=int, default=3000, help='Target amount')
args = parser.parse_args()

ticker = args.ticker
startTime = args.startTime
endTime = args.endTime
repairTime = args.repairTime
stopLoss = args.stopLoss
target = args.target

# ticker = 'NIFTY'
# startTime = '11:53:00'
# endTime = '12:00:00'
# repairTime = '11:59:00'
# stopLoss = -1500
# target = 3000

############## logging configs ##############
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s:%(name)s:%(levelname)s:%(message)s')

filename='../logs/Option_Scalper_log.txt'

#file_handler = logging.FileHandler(filename)
file_handler = TimedRotatingFileHandler(filename, when='d', interval=1, backupCount=3)
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)

stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(stream_handler)

############## XTS Initialisation ##############
cfg = configparser.ConfigParser()
cfg.read('../../XTConnect/config.ini')
source = cfg['user']['source']
appKey = cfg.get('user', 'interactive_appkey')
secretKey = cfg.get('user', 'interactive_secretkey')
xt = XTSConnect(appKey, secretKey, source)

global cdate
cdate = datetime.strftime(datetime.now(), "%d-%m-%Y")
token_file=f'access_token_{cdate}.txt'
file = Path(token_file)
if file.exists() and (date.today() == date.fromtimestamp(file.stat().st_mtime)):
    logger.info('Token file exists and created today')
    in_file = open(token_file,'r').read().split()
    access_token = in_file[0]
    userID=in_file[1]
    isInvestorClient=in_file[2]
    logger.info('Initializing session with token..')
    xt._set_common_variables(access_token, userID, isInvestorClient)
else:
    logger.error('Wrong with token file. Generate separately.. Aborting script!..')
    exit()

############## Variable Declarations ##############
logger.info(f'ticker - {ticker}')
logger.info(f'startTime - {startTime}')
logger.info(f'endTime - {endTime}')
logger.info(f'repairTime - {repairTime}')
logger.info(f'stopLoss - {stopLoss}')
logger.info(f'target - {target}')
multiplier=1
etr_inst = None
rpr_inst = None
ext_inst = None
tr_insts = None
ltp = {}
gl_pnl = None
pnl_dump = []
idxs = ['NIFTY','BANKNIFTY']
orders=[{'refId':10001, 'setno':1, 'ent_txn_type': "sell", 'rpr_txn_type': "buy", 'idx':ticker, 'otype': "ce", 'status': "Idle", 'expiry': 'week', 'lot': 1, 'startTime':startTime, 'repairEndTime':repairTime},
        {'refId':10002, 'setno':2, 'ent_txn_type': "sell", 'rpr_txn_type': "buy", 'idx':ticker, 'otype': "pe", 'status': "Idle", 'expiry': 'week', 'lot': 1, 'startTime':startTime, 'repairEndTime':repairTime}]
universal = {'exit_status': 'Idle', 'repair_status': 'not_done', 'minPrice': stopLoss, 'maxPrice': target, 'exitTime':endTime, 'ext_txn_type':'buy'}
# exitTime = datetime.strptime((cdate+" "+universal['exitTime']),"%d-%m-%Y %H:%M:%S")
# logger.info(f'args passed and orders become {orders}' )
############## Functions ##############

def get_expiry():
    global weekly_exp, monthly_exp
    now = datetime.today()
    cmon = now.month
    xpry_resp = xt.get_expiry_date(exchangeSegment=2, series='OPTIDX', symbol='NIFTY')
    if 'result' in xpry_resp:
        expiry_dates = xpry_resp['result']
    else:
        logger.error('Error getting Expiry dates..')
        raise ex.XTSDataException('Issue in getting expiry dates')

    thu = (now + relativedelta(weekday=TH(1))).strftime('%d%b%Y')
    wed = (now + relativedelta(weekday=WE(1))).strftime('%d%b%Y')

    weekly_exp = thu if thu in expiry_dates else wed
    logger.info(f'{weekly_exp} is the week expiry')

    nxtmon = (now + relativedelta(weekday=TH(1))).month
    if (nxtmon != cmon):
        month_last_thu_expiry = now + relativedelta(weekday=TH(5))
        mon_thu = (now + relativedelta(weekday=TH(5))).strftime('%d%b%Y')
        mon_wed = (now + relativedelta(weekday=WE(5))).strftime('%d%b%Y')
        if (month_last_thu_expiry.month!= nxtmon):
            mon_thu = (now + relativedelta(weekday=TH(4))).strftime('%d%b%Y')
            mon_wed = (now + relativedelta(weekday=WE(4))).strftime('%d%b%Y')
    else:
        for i in range(1, 7):
            t = now + relativedelta(weekday=TH(i))
            if t.month != cmon:
                # since t is exceeded we need last one  which we can get by subtracting -2 since it is already a Thursday.
                mon_thu = (t + relativedelta(weekday=TH(-2))).strftime('%d%b%Y')
                mon_wed = (t + relativedelta(weekday=WE(-2))).strftime('%d%b%Y')
                break
    monthly_exp = mon_thu if mon_thu in expiry_dates else mon_wed
    logger.info(f'{monthly_exp} is the month expiry')

def getSpot(idx):
    global spot
    if idx == 'NIFTY':
        ids = 'NIFTY 50'
    elif idx == 'BANKNIFTY':
        ids = 'NIFTY BANK'
    else:
        logger.info(f'Invalid Index name {idx} - Valid names are {idxs}')

    # if idx in idxs[0]:
    #     # base = 50
    #     ids = 'NIFTY 50'
    # elif idx in idxs[1]:
    #     # base = 100
    #     ids = 'NIFTY BANK'
    # else:
    #     logger.info(f'Invalid Index name {idx} - Valid names are {idxs}')
        # exit()
    try:
        idx_instruments = [{'exchangeSegment': 1, 'exchangeInstrumentID': ids}]
        spot_resp = xt.get_quote(
                    Instruments=idx_instruments,
                    xtsMessageCode=1504,
                    publishFormat='JSON')
        if spot_resp['type'] !='error':
            listQuotes = json.loads(spot_resp['result']['listQuotes'][0])
            spot = listQuotes['IndexValue']
        else:
            logger.error(spot_resp['description'])
            raise Exception()
    except Exception:
        logger.exception(f'Unable to getSpot from index {ids}')
        exit()
    else:
        # logger.info(f'spot is : {spot}')
        return spot

def strikePrice(idx, spot):
    if idx in idxs[0]:
        base = 50
    elif idx in idxs[1]:
        base = 100
    else:
        logger.info(f'Invalid Index name {idx} - Valid names are {idxs}')
    strikePrice = base * round(spot/base)
    logger.info(f'StrikePrice computed as : {strikePrice}')
    return strikePrice

def getOrderList():
    aa = 0
    logger.info('Checking OrderBook for order status..')
    while aa < 5:
        try:
           oBook_resp = xt.get_order_book()
           if oBook_resp['type'] != "error":
               orderList =  oBook_resp['result']
               # logger.info('OrderBook result retreived success')
               return orderList
               break
           else:
               raise Exception("Unkonwn error in getOrderList func")
        except Exception:
            logger.exception("Can't extract order data..retrying")
            # traceback.print_exc()
            time.sleep(2)
            aa+=1

def cancelOrder(OrderID):
    logger.info(f'Cancelling order: {OrderID} ')
    cancel_resp = xt.cancel_order(
        appOrderID=OrderID,
        orderUniqueIdentifier='FC_Cancel_Orders_1')
    if cancel_resp['type'] != 'error':
        cancelled_SL_orderID = cancel_resp['result']['AppOrderID']
        logger.info(f'Cancelled SL order id : {cancelled_SL_orderID}')
    if cancel_resp['type'] == 'error':
        logger.error(f'Cancel order not processed for : {OrderID}')

def masterDump():
    global instrument_df
    filename=f'../ohlc/NSE_Instruments_{cdate}.csv'
    file = Path(filename)
    if file.exists() and (date.today() == date.fromtimestamp(file.stat().st_mtime)):
        logger.info('MasterDump already exists.. reading directly')
        instrument_df=pd.read_csv(filename,header='infer')
    else:
        logger.info('Creating MasterDump..')
        exchangesegments = [xt.EXCHANGE_NSEFO]
        mastr_resp = xt.get_master(exchangeSegmentList=exchangesegments)
        # print("Master: " + str(mastr_resp))
        master=mastr_resp['result']
        spl=master.split('\n')
        mstr_df = pd.DataFrame([sub.split("|") for sub in spl],columns=(['ExchangeSegment','ExchangeInstrumentID','InstrumentType','Name','Description','Series','NameWithSeries','InstrumentID','PriceBand.High','PriceBand.Low','FreezeQty','TickSize',' LotSize','UnderlyingInstrumentId','UnderlyingIndexName','ContractExpiration','StrikePrice','OptionType']))
        instrument_df = mstr_df[mstr_df.Series == 'OPTIDX']
        instrument_df.to_csv(f"../ohlc/NSE_Instruments_{cdate}.csv",index=False)

def instrumentLookup(instrument_df,symbol):
    """Looks up instrument token for a given script from instrument dump"""
    try:
        return instrument_df[instrument_df.Description==symbol].ExchangeInstrumentID.values[0]
    except:
        return -1

def getLTP():
    global ltp
    # ltp={}
    if tr_insts:
        # logger.info('inside tr_insts cond - getLTP')
        symbols=[i['symbol'] for i in tr_insts if i['set_type'] == 'Entry']
        instruments=[]
        for symbol in symbols:
            instruments.append({'exchangeSegment': 2, 'exchangeInstrumentID': symbol})
        xt.send_unsubscription(Instruments=instruments,xtsMessageCode=1502)
        subs_resp=xt.send_subscription(Instruments=instruments,xtsMessageCode=1502)
        if subs_resp['type'] == 'success':
            for symbol,i in zip(symbols,range(len(symbols))):
                listQuotes = json.loads(subs_resp['result']['listQuotes'][i])
                price=listQuotes['Touchline']['LastTradedPrice']
                ltp[symbol]=price

def getGlobalPnL():
    global gl_pnl, df, gdf, pnl_dump
    # pnl_dump = []
    if tr_insts:
        # logger.info('inside tr_insts cond - getGlobalLTP')
        df = pd.DataFrame(tr_insts)
        df['tr_amount'] = df['tr_qty']*df['tradedPrice']
        df = df.fillna(0)
        df = df.astype(dtype={'set': int, 'txn_type': str, 'strike': int, 'qty': int, 'tr_qty': int, 'expiry': str, \
                                 'name': str, 'symbol': int, 'orderID': int, 'tradedPrice': float, 'dateTime': str, \
                                 'set_type': str, 'tr_amount': float, 'optionType': str})
        gdf = df.groupby(['name','symbol'],as_index=False).sum()[['symbol','name','tr_qty','tradedPrice','tr_amount']]
        gdf['ltp'] = gdf['symbol'].map(ltp)
        gdf['cur_amount'] = gdf['tr_qty']*gdf['ltp']
        gdf['pnl'] = gdf['cur_amount'] - gdf['tr_amount']
        logger.info(f'\n\nPositionList: \n {df}')
        logger.info(f'\n\nCombinedPositionsLists: \n {gdf}')
        gl_pnl = round(gdf['pnl'].sum(),2)
        logger.info(f'\n\nGlobal PnL : {gl_pnl} \n')
        pnl_dump.append([time.strftime("%d-%m-%Y %H:%M:%S"),gl_pnl])
    else:
        gl_pnl = 0

def placeOrder(symbol,txn_type,qty):
    logger.info('Placing Orders..')
    # Place an intraday stop loss order on NSE
    if txn_type == "buy":
        t_type=xt.TRANSACTION_TYPE_BUY
    elif txn_type == "sell":
        t_type=xt.TRANSACTION_TYPE_SELL
    try:
        order_resp = xt.place_order(exchangeSegment=xt.EXCHANGE_NSEFO,
                         exchangeInstrumentID= symbol ,
                         productType=xt.PRODUCT_MIS,
                         orderType=xt.ORDER_TYPE_MARKET,
                         orderSide=t_type,
                         timeInForce=xt.VALIDITY_DAY,
                         disclosedQuantity=0,
                         orderQuantity=qty,
                         limitPrice=0,
                         stopPrice=0,
                         orderUniqueIdentifier="FC_MarketOrder"
                         )
        if order_resp['type'] != 'error':
            orderID = order_resp['result']['AppOrderID']            #extracting the order id from response
            logger.info(f'Order ID for {t_type} {symbol} is: {orderID}')
            time.sleep(2)
            a=0
            while a<12:
                orderLists = getOrderList()
                if orderLists:
                    new_orders = [ol for ol in orderLists if ol['AppOrderID'] == orderID and ol['OrderStatus'] != 'Filled']
                    if not new_orders:
                        tradedPrice = float(next((orderList['OrderAverageTradedPrice'] \
                                            for orderList in orderLists \
                                                if orderList['AppOrderID'] == orderID and \
                                                    orderList['OrderStatus'] == 'Filled'),None).replace(',', ''))
                        LastUpdateDateTime=datetime.fromisoformat(next((orderList['LastUpdateDateTime'] for orderList in orderLists if orderList['AppOrderID'] == orderID and orderList['OrderStatus'] == 'Filled'))[0:19])
                        dateTime = LastUpdateDateTime.strftime("%Y-%m-%d %H:%M:%S")
                        logger.info(f"traded price is: {tradedPrice} and ordered  time is: {dateTime}")
                        return orderID, tradedPrice, dateTime
                        break
                        # loop = False
                    else:
                        logger.info(f' Placed order {orderID} might be in Open or New Status, Hence retrying..{a}')
                        a+=1
                        time.sleep(2.5)
                        if a==11:
                            logger.info('Placed order is still in New or Open Status..Hence Cancelling the placed order')
                            cancelOrder(orderID)
                            return None, None, None
                            break
                else:
                    logger.info('\n  Unable to get OrderList inside place order function..')
                    logger.info('..Hence traded price will retun as Zero \n ')
        elif order_resp['type'] == 'error':
            logger.error(order_resp['description'])
            logger.info(f'Order not placed for - {symbol} ')
            raise Exception('Order not placed - ')
    except Exception():
        raise ex.XTSOrderException('Unable to place order in placeOrder func...')
        logger.exception('Unable to place order in placeOrder func...')
        time.sleep(1)

def execute(orders):
    global tr_insts, spot
    tr_insts = []
    etr_inst = {}
    rpr_inst = {}
    rpr_inst2 = {}
    startTime = datetime.strptime((cdate+" "+orders['startTime']),"%d-%m-%Y %H:%M:%S")
    weekday = datetime.today().weekday()

    orders['straddle_points'] = 50 if orders['idx'] == 'NIFTY' else 100
    orders['ptsChg'] = 40 if orders['idx'] == 'NIFTY' else 90
    logger.info(f'orders after spoints and ptschg: {orders}')
    while True:
        time.sleep(1)
        if orders['status'] == 'Idle':
            #Entry condition check
            if (datetime.now() >= startTime):
                logger.info(f'Placing orders for {orders["setno"]} at {orders["startTime"]}..')
                etr_inst['set'] = orders['setno']
                etr_inst['txn_type'] = orders['ent_txn_type']
                etr_inst['qty'] = 75*orders['lot'] if orders['idx'] == 'NIFTY' else 25*orders['lot']
                etr_inst['spot'] = getSpot(orders['idx'])
                sp = strikePrice(orders['idx'], etr_inst['spot'])
                if weekday != 3: #if not thursday, take straddle
                    etr_inst['strike'] = sp + orders['straddle_points'] if orders['otype'] == 'ce' else sp - orders['straddle_points']
                else:
                    etr_inst['strike'] = sp
                etr_inst['tr_qty'] = -etr_inst['qty'] if orders['ent_txn_type'] == 'sell' else etr_inst['qty']

                if orders['expiry'] == 'week':
                    etr_inst['expiry'] = weekly_exp
                etr_inst['optionType'] = orders['otype'].upper()

                if weekly_exp == monthly_exp:
                    inst_name = orders['idx']+(datetime.strftime(datetime.strptime(etr_inst['expiry'], '%d%b%Y'),'%y%b')).upper()+str(etr_inst['strike'])+etr_inst['optionType']
                else:
                    inst_name = orders['idx']+(datetime.strftime(datetime.strptime(etr_inst['expiry'], '%d%b%Y'),'%y%#m%d'))+str(etr_inst['strike'])+etr_inst['optionType']
                etr_inst['name'] = inst_name
                etr_inst['symbol'] = int(instrumentLookup(instrument_df,inst_name))
                etr_inst['orderID'] = None
                etr_inst['tradedPrice'] = None
                # logger.info(f"orders before placing orders : {orders}")
                orderID, tradedPrice, dateTime = placeOrder(etr_inst['symbol'],etr_inst['txn_type'],etr_inst['qty'])
                etr_inst['orderID'] = orderID
                etr_inst['tradedPrice'] = tradedPrice
                etr_inst['dateTime'] = dateTime
                if orderID and tradedPrice:
                    etr_inst['set_type'] = 'Entry'
                    orders['status'] = 'Entered'
                logger.info(f'Entry order dtls of {etr_inst["set"]}: {etr_inst}')
                tr_insts.append(etr_inst)

        if universal['exit_status'] == 'Idle' and universal['repair_status'] == 'not_done':  #Checking wheather universal exit triggered or not
            if orders['status'] == 'Entered':
                repairEndTime = datetime.strptime((cdate+" "+orders['repairEndTime']),"%d-%m-%Y %H:%M:%S")
                if datetime.now() <= repairEndTime:
                    ltpOfIdx = spot
                    ptsChng = orders['ptsChg'] if orders['otype']  == 'ce' else -orders['ptsChg']
                    # logger.info(f'Points difference in {orders["otype"]} : {round((ltpOfIdx - etr_inst["spot"]),2)}')
                    #repair condition check
                    if (ltpOfIdx > (etr_inst['spot'] +  ptsChng) and orders['otype']  == 'ce')\
                        or (ltpOfIdx < (etr_inst['spot'] +  ptsChng) and orders['otype']  == 'pe'):
                        logger.info(f'Points difference in {orders["otype"]} : {round((ltpOfIdx - etr_inst["spot"]),2)}')
                        logger.info(f'Repair condition met in set: {orders["setno"]}..')
                        #buy what sold in entry
                        rpr_inst['set']=orders['setno']
                        rpr_inst['txn_type'] = orders['rpr_txn_type']
                        rpr_inst['strike'] = etr_inst['strike']
                        rpr_inst['qty'] = etr_inst['qty']
                        rpr_inst['tr_qty'] = -rpr_inst['qty'] if orders['rpr_txn_type'] == 'sell' else rpr_inst['qty']
                        if orders['expiry'] == 'week':
                            rpr_inst['expiry'] = weekly_exp
                        rpr_inst['optionType'] = orders['otype'].upper()
                        rpr_inst['name'] = etr_inst['name']
                        rpr_inst['symbol'] = etr_inst['symbol']
                        rpr_inst['orderID'] = None
                        rpr_inst['tradedPrice'] = None
                        orderID, tradedPrice, dateTime = placeOrder(rpr_inst['symbol'],rpr_inst['txn_type'],rpr_inst['qty'])
                        rpr_inst['orderID'] = orderID
                        rpr_inst['tradedPrice'] = tradedPrice
                        rpr_inst['dateTime'] = dateTime
                        if orderID and tradedPrice:
                            # orders['status'] = '1Repaired'
                            rpr_inst['set_type'] = 'Repair'
                        logger.info(f'Repair1 order dtls of set {rpr_inst["set"]}: {rpr_inst}')
                        tr_insts.append(rpr_inst)

                        # second repair
                        if rpr_inst['set_type'] == 'Repair':
                            logger.info(f'Second Repair condition met in set: {orders["setno"]}..')
                            rpr_inst2['set'] = orders['setno']
                            rpr_inst2['txn_type'] = orders['ent_txn_type']
                            rpr_inst2['strike'] = etr_inst['strike'] + orders['straddle_points'] \
                                                    if orders['otype']  == 'ce' \
                                                        else etr_inst['strike'] - orders['straddle_points']
                            rpr_inst2['qty'] = etr_inst['qty']
                            rpr_inst2['tr_qty'] = -rpr_inst2['qty'] if orders['ent_txn_type'] == 'sell' else rpr_inst2['qty']
                            if orders['expiry'] == 'week':
                                rpr_inst2['expiry'] = weekly_exp
                            rpr_inst2['optionType'] = orders['otype'].upper()

                            if weekly_exp == monthly_exp:
                                inst_name = orders['idx']+(datetime.strftime(datetime.strptime(rpr_inst2['expiry'], '%d%b%Y'),'%y%b')).upper()+str(rpr_inst2['strike'])+rpr_inst2['optionType']
                            else:
                                inst_name = orders['idx']+(datetime.strftime(datetime.strptime(rpr_inst2['expiry'], '%d%b%Y'),'%y%#m%d'))+str(rpr_inst2['strike'])+rpr_inst2['optionType']
                            rpr_inst2['name'] = inst_name
                            rpr_inst2['symbol'] = int(instrumentLookup(instrument_df, rpr_inst2['name']))
                            rpr_inst2['orderID'] = None
                            rpr_inst2['tradedPrice'] = None
                            orderID, tradedPrice, dateTime = placeOrder(rpr_inst2['symbol'],rpr_inst2['txn_type'],rpr_inst2['qty'])
                            rpr_inst2['orderID'] = orderID
                            rpr_inst2['tradedPrice'] = tradedPrice
                            rpr_inst2['dateTime'] = dateTime
                            if orderID and tradedPrice:
                                orders['status'] = 'Repaired'
                                rpr_inst2['set_type'] = 'Entry'
                            logger.info(f'Repair2 order dtls of {rpr_inst2["set"]}: {rpr_inst2}')
                            tr_insts.append(rpr_inst2)
                            universal['repair_status'] = 'Done'
                            continue

        elif universal['exit_status'] == 'Exited':
           orders['status'] = 'Universal_Exit'
           logger.info('Orders must be square-off by Universal Exit Func')
           break
        #breaking set loop if status repaired
        if orders['status'] == 'Repaired':
            logger.info(f'Repaired the Entry Order set: {orders["setno"]}. Exiting the thread')
            break

def exitCheck(universal):
    global tr_insts #todo add gl_pnl to global if the conditions didnt work
    # pnl_dump=[]
    ext_inst = {}
    exitTime = datetime.strptime((cdate+" "+universal['exitTime']),"%d-%m-%Y %H:%M:%S")
    # print('exitTime:', exitTime)
    while True:
        if universal['exit_status'] == 'Idle':
            #Exit condition check
            # logger.info(f'exitcheck - {gl_pnl}') #todo comment this line after execution
            if (datetime.now() >= exitTime) or (gl_pnl <= universal['minPrice']) or (gl_pnl >= universal['maxPrice']):
                logger.info('Exit time condition passed. Squaring off all open positions')
                for i in range(len(gdf)):
                    if gdf["tr_qty"].values[i] == 0:
                        continue
                    ext_inst['symbol'] = int(gdf['symbol'].values[i])
                    ext_inst['qty'] = abs(int(gdf['tr_qty'].values[i]))
                    ext_inst['tr_qty'] = -ext_inst['qty'] if universal['ext_txn_type'] == 'sell' else ext_inst['qty']
                    ext_inst['txn_type'] = universal['ext_txn_type']
                    ext_inst['name'] = str(gdf['name'].values[i])
                    ext_inst['optionType'] = ext_inst['name'][-2:]
                    ext_inst['strike'] = ext_inst['name'][-7:-2]
                    ext_inst['orderID'] = None
                    ext_inst['tradedPrice'] = None
                    orderID, tradedPrice, dateTime = placeOrder(ext_inst['symbol'], ext_inst['txn_type'], ext_inst['qty'])
                    ext_inst['orderID'] = orderID
                    ext_inst['tradedPrice'] = tradedPrice
                    ext_inst['dateTime'] = dateTime
                    if orderID and tradedPrice:
                        ext_inst['set_type']='Universal_Exit'
                        # universal['exit_status'] = 'Exited'
                    logger.info(f'Universal Exit order dtls: {ext_inst}')
                    tr_insts.append(ext_inst.copy())
                logger.info('Breaking the main loop')
                universal['exit_status'] = 'Exited'
                break
            else:
                time.sleep(1)

def dataToExcel(pnl_dump):
    blueFill = PatternFill(start_color='000000FF', end_color='000000FF',
                   fill_type='solid')
    time.sleep(randint(3,9))
    filename = os.path.basename(__file__).split('.')[0]
    sheetname = cdate+'_'+startTime.replace(':','_')
    pnl_df = pd.DataFrame(pnl_dump,columns=['date','pl'])
    pnl_df = pnl_df.set_index(['date'])
    pnl_df.index = pd.to_datetime(pnl_df.index, format='%Y-%m-%d %H:%M:%S')
    resampled_df = pnl_df['pl'].resample('1min').ohlc()
    #writing the output to excel sheet
    writer = pd.ExcelWriter(f'..\\pnl\\{filename}.xlsx',engine='openpyxl')
    writer.book = load_workbook(f'..\\pnl\\{filename}.xlsx')
    resampled_df.to_excel(writer, sheet_name=(sheetname), index=True)
    df.to_excel(writer, sheet_name=(sheetname),startrow=11, startcol=6, index=False)
    gdf.to_excel(writer, sheet_name=(sheetname),startrow=5, startcol=6, index=False)
    writer.sheets=dict((ws.title, ws) for ws in writer.book.worksheets)
    worksheet = writer.sheets[sheetname]
    worksheet['G1'] = f"{filename} - {sheetname}"
    worksheet['G1'].font = Font(bold=True)
    worksheet['G1'].fill = blueFill
    worksheet['G3'] = "MaxPnL"
    worksheet['G3'].font = Font(bold=True)
    worksheet["G4"] = "=MAX(E:E)"
    worksheet['H3'] = "MinPnL"
    worksheet['H3'].font = Font(bold=True)
    worksheet["H4"] = "=MIN(E:E)"
    worksheet['I3'] = "FinalPnL"
    worksheet['I3'].font = Font(bold=True)
    worksheet['I4'] = gl_pnl
    writer.save()
    writer.close()

############## main ##############

if __name__ == '__main__':
    threads=[]
    try:
        masterDump()
        get_expiry()
    except:
        logger.exception('Failed to get masterDump/ expiryDates. Exiting..')
        exit()
    # all the sets will execute in parallel with threads
    for i in range(len(orders)):
        t = Thread(target=execute,args=(orders[i],))
        t.start()
        threads.append(t)

    # below function runs in background
    logger.info('Starting a timer based thread to fetch LTP of traded instruments..')
    getGlobalPnL()
    fetchSpot = timer.RepeatedTimer(8, getSpot,ticker)
    fetchLtp = timer.RepeatedTimer(9, getLTP)
    fetchPnL = timer.RepeatedTimer(10, getGlobalPnL)
    try:
        exitCheck(universal)
        time.sleep(5)
    except KeyboardInterrupt:
        logger.error('\n\nKeyboard exception received. Exiting.')
        universal['exit_status'] = 'Exited' #todo write dead case here to stop the threads in case of exitcheck exception
        exit()
    except Exception:
        universal['exit_status'] = 'Exited'
        logger.exception('Error Occured..')
    finally:
        logger.info('Cleaning up..')
        fetchSpot.stop()
        fetchLtp.stop()
        fetchPnL.stop()
        # _ = [t.join() for t in threads]
        t.join()
        time.sleep(5)
        #prints dump to excel
        getGlobalPnL()      #getting latest data
        dataToExcel(pnl_dump)
        # logging the orders and data to log file
        logger.info('--------------------------------------------')
        logger.info(f'Total Orders and its status: \n {tr_insts} \n')
        logger.info('********** Summary **********')
        logger.info(f'\n\n PositionList: \n {df}')
        logger.info(f'\n\n CombinedPositionsLists: \n {gdf}')
        logger.info(f'\n\n Global PnL : {gl_pnl} \n')
        logger.info('--------------------------------------------')

############# END ##############

# !python ./OptionScalper.py -t NIFTY -s 17:33:00
