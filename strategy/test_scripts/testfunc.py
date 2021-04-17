# -*- coding: utf-8 -*-
"""
Created on Sat Apr 17 16:06:13 2021

@author: lmahendran
"""
import pandas as pd
from datetime import datetime,date
from dateutil.relativedelta import relativedelta, TH, WE
from XTConnect.Connect import XTSConnect
import XTConnect.Exception as ex
from pathlib import Path
import time
import json
import logging
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
# pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', -1)
import configparser
import argparse
import timer
from threading import Thread
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill
from logging.handlers import TimedRotatingFileHandler
from sys import exit
import os
from random import randint

pnl_dump=[('2021-04-15 13:31:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:33:00',123),('2021-04-15 13:30:10',123),('2021-04-15 13:30:20',123),('2021-04-15 13:30:50',123),('2021-04-15 13:30:32',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123),('2021-04-15 13:30:00',123)]
cdate = datetime.strftime(datetime.now(), "%d-%m-%Y")
startTime = '1:05:00'



data = {
  "calories": [420, 380, 390],
  "duration": [50, 40, 45]
}
df = pd.DataFrame(data)
gdf = pd.DataFrame(data)

gl_pnl=1000000000

def dataToExcel(pnl_dump):
    redFill = PatternFill(start_color='000000FF',
                   end_color='000000FF',
                   fill_type='solid')
    time.sleep(randint(3,9))
    filename = os.path.basename(__file__).split('.')[0]
    sheetname = cdate+'_'+startTime.replace(':','_')
    pnl_df = pd.DataFrame(pnl_dump,columns=['date','pl'])
    pnl_df = pnl_df.set_index(['date'])
    pnl_df.index = pd.to_datetime(pnl_df.index, format='%Y-%m-%d %H:%M:%S')
    resampled_df = pnl_df['pl'].resample('1min').ohlc()
    #writing the output to excel sheet
    writer = pd.ExcelWriter(f'..\\pnl\\{filename}.xlsx',engine='openpyxl')
    writer.book = load_workbook(f'..\\pnl\\{filename}.xlsx')
    resampled_df.to_excel(writer, sheet_name=(sheetname), index=True)
    df.to_excel(writer, sheet_name=(sheetname),startrow=11, startcol=6, index=False)
    gdf.to_excel(writer, sheet_name=(sheetname),startrow=4, startcol=6, index=False)
    writer.sheets=dict((ws.title, ws) for ws in writer.book.worksheets)
    worksheet = writer.sheets[sheetname]
    worksheet['G1'] = f"{filename} - {sheetname}"
    worksheet['G1'].font = Font(bold=True)
    worksheet['G1'].fill = redFill
    worksheet['G2'] = "MaxPnL"
    worksheet['G2'].font = Font(bold=True)
    worksheet["G3"] = "=MAX(E:E)"
    worksheet['H2'] = "MinPnL"
    worksheet['H2'].font = Font(bold=True)
    worksheet["H3"] = "=MIN(E:E)"
    worksheet['I2'] = "FinalPnL"
    worksheet['I2'].font = Font(bold=True)
    worksheet['I3'] = gl_pnl
    writer.save()
    writer.close()
    
if __name__ == '__main__':
    dataToExcel(pnl_dump)